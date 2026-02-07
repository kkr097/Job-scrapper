"""
Excel Exporter - Exports jobs to a formatted, color-coded Excel file.
"""

from datetime import datetime, timedelta
import os
from typing import List, Dict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import json


def export_to_excel(jobs: List[Dict], output_path: str = "daily_jobs.xlsx"):
    """
    Export jobs to an Excel file with color coding.
    
    Colors:
    - Green (8-10): Great match
    - Yellow (5-7): Okay match
    - Red (1-4): Poor match
    """
    
    # Sort by score (highest first)
    jobs_sorted = sorted(jobs, key=lambda x: x.get('score', 0), reverse=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Matches"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Score", "Title", "Company", "Location", "Source", "Match Reasons", "Missing", "URL", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Column widths
    widths = [8, 45, 25, 20, 12, 40, 30, 50, 70]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Data rows
    for row, job in enumerate(jobs_sorted, 2):
        score = job.get('score', 0)
        
        # Determine row color
        if score >= 8:
            row_fill = green_fill
        elif score >= 5:
            row_fill = yellow_fill
        else:
            row_fill = red_fill
        
        # Write cells
        values = [
            score,
            job.get('title', ''),
            job.get('company', ''),
            job.get('location', ''),
            job.get('source', ''),
            job.get('match_reasons', ''),
            job.get('missing_skills', ''),
            job.get('url', ''),
            job.get('description', '')
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Make URL clickable
            if col == 8 and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add summary sheet
    summary = wb.create_sheet("Summary")
    summary['A1'] = "Job Search Summary"
    summary['A1'].font = Font(bold=True, size=14)
    
    summary['A3'] = "Date:"
    summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    summary['A4'] = "Total Jobs:"
    summary['B4'] = len(jobs)
    
    summary['A5'] = "Great Matches (8-10):"
    summary['B5'] = len([j for j in jobs if j.get('score', 0) >= 8])
    summary['B5'].fill = green_fill
    
    summary['A6'] = "Good Matches (5-7):"
    summary['B6'] = len([j for j in jobs if 5 <= j.get('score', 0) < 8])
    summary['B6'].fill = yellow_fill
    
    summary['A7'] = "Poor Matches (1-4):"
    summary['B7'] = len([j for j in jobs if j.get('score', 0) < 5])
    summary['B7'].fill = red_fill
    
    # Save
    wb.save(output_path)
    print(f"✓ Saved to: {output_path}")
    
    return output_path


def append_new_to_excel(jobs: List[Dict], output_path: str = "daily_jobs.xlsx"):
    """
    Append new jobs to an existing Excel file, skipping duplicates by Title+Company+URL.
    If file doesn't exist, create it.
    """
    if not jobs:
        print("✓ No new jobs to append.")
        return output_path

    if not os.path.exists(output_path):
        return export_to_excel(jobs, output_path)

    wb = load_workbook(output_path)
    ws = wb["Job Matches"] if "Job Matches" in wb.sheetnames else wb.active

    # Find header indexes
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers) if h}
    title_col = col_map.get("Title", 2)
    company_col = col_map.get("Company", 3)
    url_col = col_map.get("URL", 8)
    score_col = col_map.get("Score", 1)
    source_col = col_map.get("Source", 5)
    location_col = col_map.get("Location", 4)
    reasons_col = col_map.get("Match Reasons", 6)
    missing_col = col_map.get("Missing", 7)

    # Build existing keys (URL only)
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = (row[url_col - 1] or "").strip().lower()
        if url:
            existing.add(url)

    # Styles
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    new_count = 0
    for job in jobs:
        title = (job.get("title", "") or "").strip()
        company = (job.get("company", "") or "").strip()
        url = (job.get("url", "") or "").strip()
        key = url.lower()
        if key in existing:
            continue

        next_row = ws.max_row + 1
        score = job.get("score", 0)
        if score >= 8:
            row_fill = green_fill
        elif score >= 5:
            row_fill = yellow_fill
        else:
            row_fill = red_fill

        values = {
            score_col: score,
            title_col: title,
            company_col: company,
            location_col: job.get("location", ""),
            source_col: job.get("source", ""),
            reasons_col: job.get("match_reasons", ""),
            missing_col: job.get("missing_skills", ""),
            url_col: url
        }

        for col, value in values.items():
            cell = ws.cell(row=next_row, column=col, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col == url_col and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

        new_count += 1
        existing.add(key)

    # Update summary sheet if present
    if "Summary" in wb.sheetnames:
        summary = wb["Summary"]
        # Recompute totals from sheet
        scores = [r[score_col - 1] for r in ws.iter_rows(min_row=2, values_only=True) if r]
        summary['B4'] = len(scores)
        summary['B5'] = len([s for s in scores if isinstance(s, (int, float)) and s >= 8])
        summary['B6'] = len([s for s in scores if isinstance(s, (int, float)) and 5 <= s < 8])
        summary['B7'] = len([s for s in scores if isinstance(s, (int, float)) and s < 5])

    wb.save(output_path)
    print(f"✓ Appended {new_count} new jobs to: {output_path}")
    return output_path


def append_new_to_csv(jobs: List[Dict], output_path: str = "daily_jobs.csv"):
    """
    Append new jobs to a CSV, skipping duplicates by URL only.
    If file doesn't exist, create it with headers.
    """
    if not jobs:
        print("✓ No new jobs to append.")
        return output_path

    headers = ["Score", "Title", "Company", "Location", "Source", "Match Reasons", "Missing", "URL", "Description"]
    existing = set()

    if os.path.exists(output_path):
        with open(output_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                url = (row.get("URL") or "").strip().lower()
                if url:
                    existing.add(url)

    new_rows = []
    for job in jobs:
        url = (job.get("url", "") or "").strip()
        if url.lower() in existing:
            continue
        new_rows.append({
            "Score": job.get("score", 0),
            "Title": job.get("title", ""),
            "Company": job.get("company", ""),
            "Location": job.get("location", ""),
            "Source": job.get("source", ""),
            "Match Reasons": job.get("match_reasons", ""),
            "Missing": job.get("missing_skills", ""),
            "URL": url,
            "Description": job.get("description", "")
        })
        if url:
            existing.add(url.lower())

    if not os.path.exists(output_path):
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(new_rows)
    else:
        with open(output_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writerows(new_rows)

    print(f"✓ Appended {len(new_rows)} new jobs to: {output_path}")
    return output_path


_WARNED_CSV = set()
_ALLOW_EMPTY_CSV = {"daily_jobs.csv", "daily_jobs_nonmatch.csv"}


def load_existing_urls(csv_path: str) -> set:
    """Load existing job URLs from a CSV file."""
    existing = set()
    if not os.path.exists(csv_path):
        if csv_path not in _WARNED_CSV:
            print(f"   ⚠ CSV not found: {csv_path}")
            _WARNED_CSV.add(csv_path)
        return existing
    row_count = 0
    url_count = 0
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            row_count += 1
            url = (row.get("URL") or row.get("url") or "").strip().lower()
            if url:
                url_count += 1
                existing.add(url)
    base_name = os.path.basename(csv_path)
    if (row_count == 0 or url_count == 0) and csv_path not in _WARNED_CSV:
        if base_name not in _ALLOW_EMPTY_CSV:
            print(f"   ⚠ CSV empty or missing URLs: {csv_path}")
        _WARNED_CSV.add(csv_path)
    return existing


def load_cache(cache_path: str) -> dict:
    """Load seen jobs cache from JSON."""
    if not os.path.exists(cache_path):
        return {}
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_cache(cache_path: str, cache: dict) -> None:
    """Save seen jobs cache to JSON."""
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def prune_cache(cache: dict, max_days: int) -> dict:
    """Remove cache entries older than max_days based on last_seen."""
    if not max_days:
        return cache
    cutoff = datetime.now() - timedelta(days=max_days)
    pruned = {}
    for url, entry in cache.items():
        last_seen = entry.get("last_seen")
        if not last_seen:
            pruned[url] = entry
            continue
        try:
            ts = datetime.fromisoformat(last_seen)
            if ts >= cutoff:
                pruned[url] = entry
        except Exception:
            pruned[url] = entry
    return pruned


if __name__ == "__main__":
    # Test with sample data
    sample_jobs = [
        {
            'title': 'Mechanical Design Engineer',
            'company': 'BMW Group',
            'location': 'Munich, Germany',
            'source': 'LinkedIn',
            'url': 'https://linkedin.com/jobs/123',
            'score': 9,
            'match_reasons': 'CATIA V5, automotive, design',
            'missing_skills': 'German B2'
        },
        {
            'title': 'CAD Engineer',
            'company': 'Siemens',
            'location': 'Berlin, Germany',
            'source': 'Indeed',
            'url': 'https://indeed.com/jobs/456',
            'score': 7,
            'match_reasons': 'CAD design experience',
            'missing_skills': 'NX experience'
        },
        {
            'title': 'Software Engineer',
            'company': 'Google',
            'location': 'Munich, Germany',
            'source': 'StepStone',
            'url': 'https://stepstone.de/jobs/789',
            'score': 2,
            'match_reasons': 'Wrong field',
            'missing_skills': 'Programming skills'
        }
    ]
    
    export_to_excel(sample_jobs, "test_jobs.xlsx")
    print("Test file created: test_jobs.xlsx")
